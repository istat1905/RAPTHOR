import pandas as pd
from openpyxl import load_workbook

def appliquer_criteres(df, fichier_modele="modele.xlsx", fichier_sortie="commandes_final.xlsx"):
    # Exemple : filtrer par client Auchan France
    df_filtre = df[df[2] == "Auchan France"]

    # Charger modèle Excel et écrire les données
    wb = load_workbook(fichier_modele)
    ws = wb.active
    for r_idx, row in enumerate(df_filtre.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(fichier_sortie)
    return fichier_sortie
